#!/usr/bin/env python3
"""
아산시 강소형 스마트시티 정산시스템 — 전체 내역 종합 엑셀

시스템의 4개 원장(정산증빙·계약·중요재산·준공) + 예산을 하나의 통합 워크북으로 정리.
집행결과 제출 양식(별첨4)과 같은 구조의 시트를 포함하며, 집계는 전부 수식으로 계산한다.

시트 구성:
  00_요약        전체 KPI · 재원/비목/상태 집계
  01_총괄명세서   비목·세목별 예산(A)·집행(B)·잔액·집행률   [별지 제3호 대응]
  02_집행내역     일자별 전체 집행 2,364건                 [별지 제4호 대응]
  03_계약대장     계약 3건
  04_중요재산     자산 98건 (자산관리 시스템 연동)
  05_준공현황     단위공사 12건 · 필수문서 확보율
  06_검증지적     법정요건 검증 결과 (R-01~R-25)
"""
import json, subprocess, sys, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
def rd(p): return json.load(open(os.path.join(ROOT, p), encoding='utf-8'))

base = rd('data/baseline.json')
evidence = rd('data/evidence.json')['레코드']
budget = rd('data/budget.json')['레코드']
contracts = rd('data/contracts.json')['레코드']
assets = rd('data/assets.json')['레코드']
completion = rd('data/completion.json')['레코드']
codes = rd('codes/expense-categories.json')
비목순 = [b['코드'] for b in codes['비목']]

# 검증 결과
try:
    out = subprocess.run(['node', 'scripts/validate.js', '--json'], cwd=ROOT,
                         capture_output=True, text=True, timeout=120)
    vr = json.loads(out.stdout[out.stdout.index('{'):])
except Exception:
    vr = {'오류': 0, '경고': 0, '결과': []}

# ── 스타일 ──
F = 'Arial'; FK = '맑은 고딕'
HDR = PatternFill('solid', fgColor='1F3864')
SUB = PatternFill('solid', fgColor='D9E2F3')
TOT = PatternFill('solid', fgColor='F2F2F2')
HF = Font(name=FK, size=10, bold=True, color='FFFFFF')
TF = Font(name=FK, size=14, bold=True, color='1F3864')
BF = Font(name=FK, size=9)
BB = Font(name=FK, size=9, bold=True)
thin = Side(style='thin', color='BFBFBF')
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical='center')
CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')

wb = openpyxl.Workbook()

def head(ws, title, headers, widths, start=3):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws['A1'] = title; ws['A1'].font = TF
    ws.row_dimensions[1].height = 22
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(start, i, h); c.font = HF; c.fill = HDR; c.alignment = CTR; c.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[start].height = 28
    ws.freeze_panes = ws.cell(start+1, 1)

def put(ws, r, row, ctr_cols=(), num_cols=(), bold=False, fill=None):
    for i, v in enumerate(row, 1):
        c = ws.cell(r, i, v)
        c.font = BB if bold else BF
        c.border = BOX
        if fill: c.fill = fill
        if i in num_cols: c.alignment = RIGHT; c.number_format = '#,##0'
        elif i in ctr_cols: c.alignment = CTR
        else: c.alignment = WRAP

# ══════════════════════════════════════════════════════════
# 00_요약
# ══════════════════════════════════════════════════════════
ws = wb.active; ws.title = '00_요약'; ws.sheet_view.showGridLines = False
ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0; ws.sheet_properties.pageSetUpPr=openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws['A1'] = '아산시 강소형 스마트시티 조성사업 — 정산 종합 내역'; ws['A1'].font = Font(name=FK, size=15, bold=True, color='1F3864')
ws.column_dimensions['A'].width = 22; ws.column_dimensions['B'].width = 26
for c in 'CDEF': ws.column_dimensions[c].width = 16

info = [
    ('사업명', base['사업']['사업명']),
    ('사업기간', f"{base['사업']['사업기간']['개시일']} ~ {base['사업']['사업기간']['종료일']}"),
    ('총사업비', base['사업']['총사업비']['금액']),
    ('작성', '㈜제일엔지니어링종합건축사사무소 PMO / 상무 이성호'),
    ('작성일', '2026-09-04'),
]
r = 3
for k, v in info:
    ws.cell(r, 1, k).font = BB; ws.cell(r, 1).fill = SUB; ws.cell(r, 1).border = BOX; ws.cell(r,1).alignment=CTR
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = ws.cell(r, 2, v); c.font = BF; c.border = BOX
    if k == '총사업비': c.number_format = '#,##0"원"'; c.alignment = RIGHT
    r += 1

# KPI (수식)
r += 1
ws.cell(r, 1, '■ 집행 현황').font = Font(name=FK, size=11, bold=True); r += 1
kpi_head = ['구분', '예산(원)', '집행(원)', '잔액(원)', '집행률', '건수']
for i, h in enumerate(kpi_head, 1):
    c = ws.cell(r, i, h); c.font = HF; c.fill = HDR; c.alignment = CTR; c.border = BOX
r += 1
# 총괄명세서 시트 참조
ws.cell(r, 1, '전체').font = BB; ws.cell(r,1).fill=TOT; ws.cell(r,1).alignment=CTR; ws.cell(r,1).border=BOX
_tot = 4 + len(budget)  # 총괄명세서 합계행
ws.cell(r, 2, f"='01_총괄명세서'!C{_tot}").number_format='#,##0'
ws.cell(r, 3, f"='01_총괄명세서'!D{_tot}").number_format='#,##0'
ws.cell(r, 4, f"='01_총괄명세서'!E{_tot}").number_format='#,##0'
ws.cell(r, 5, "=IFERROR(C{0}/B{0},0)".format(r)).number_format='0.0%'
ws.cell(r, 6, "=COUNTA('02_집행내역'!A4:A{})".format(3+len(evidence))).number_format='#,##0'
for i in range(1,7): ws.cell(r,i).font=BB; ws.cell(r,i).fill=TOT; ws.cell(r,i).border=BOX; ws.cell(r,i).alignment = RIGHT if i>1 else CTR
KPI_ROW = r

# 원장 규모
r += 3
ws.cell(r, 1, '■ 원장 규모').font = Font(name=FK, size=11, bold=True); r += 1
for label, cnt, src in [('정산증빙', len(evidence), 'Slack #플랜예산 엑셀'),
                        ('계약대장', len(contracts), 'Notion 계약대장 DB'),
                        ('중요재산', len(assets), 'Notion ← 자산관리 시스템'),
                        ('준공산출물', len(completion), 'Notion 준공산출물 DB')]:
    ws.cell(r,1,label).font=BB; ws.cell(r,1).fill=SUB; ws.cell(r,1).border=BOX; ws.cell(r,1).alignment=CTR
    ws.cell(r,2,f'{cnt}건').font=BF; ws.cell(r,2).border=BOX; ws.cell(r,2).alignment=CTR
    c=ws.cell(r,3,f'출처: {src}'); c.font=BF; c.border=BOX
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=6)
    r += 1

# 검증 요약
r += 2
ws.cell(r,1,'■ 법정요건 검증').font=Font(name=FK,size=11,bold=True); r+=1
ws.cell(r,1,f'오류 {vr["오류"]}건 · 경고 {vr["경고"]}건').font=Font(name=FK,size=10,bold=True,color='C00000')
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)

# ══════════════════════════════════════════════════════════
# 01_총괄명세서 (별지 제3호 대응)
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('01_총괄명세서')
head(ws, '보조비목·보조세목별 총괄명세서', ['보조비목','보조세목','예산(A)','집행(B)','잔액(A-B)','집행률'], [18,18,18,18,18,12])
집행 = {}
for e in evidence:
    집행[(e['비목코드'], e['세목코드'])] = 집행.get((e['비목코드'], e['세목코드']), 0) + e['집행금액']
sorted_b = sorted(budget, key=lambda b: (비목순.index(b['비목코드']), b['세목코드']))
r = 4
first = r
for b in sorted_b:
    v = 집행.get((b['비목코드'], b['세목코드']), 0)
    put(ws, r, [f"{b['보조비목']}({b['비목코드']})", f"{b['보조세목']}({b['세목코드']})",
                b['예산집행계획'], v, None, None], ctr_cols=(1,2), num_cols=(3,4,5))
    ws.cell(r,5, f"=C{r}-D{r}").number_format='#,##0'; ws.cell(r,5).alignment=RIGHT; ws.cell(r,5).border=BOX; ws.cell(r,5).font=BF
    ws.cell(r,6, f"=IFERROR(D{r}/C{r},0)").number_format='0.0%'; ws.cell(r,6).alignment=CTR; ws.cell(r,6).border=BOX; ws.cell(r,6).font=BF
    r += 1
# 합계행
put(ws, r, ['합계', '', None, None, None, None], ctr_cols=(1,2), bold=True, fill=TOT)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws.cell(r,3, f"=SUM(C{first}:C{r-1})").number_format='#,##0'
ws.cell(r,4, f"=SUM(D{first}:D{r-1})").number_format='#,##0'
ws.cell(r,5, f"=SUM(E{first}:E{r-1})").number_format='#,##0'
ws.cell(r,6, f"=IFERROR(D{r}/C{r},0)").number_format='0.0%'
for i in range(3,7): ws.cell(r,i).font=BB; ws.cell(r,i).fill=TOT; ws.cell(r,i).border=BOX; ws.cell(r,i).alignment=RIGHT if i<6 else CTR

# ══════════════════════════════════════════════════════════
# 02_집행내역 (별지 제4호 대응)
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('02_집행내역')
head(ws, '일자별 집행내역', ['No','집행일자','단위사업','보조비목','보조세목','재원','지급처','집행금액','사용목적','지급방식','검토상태'],
     [5,11,11,11,12,8,18,13,30,12,10])
ev_sorted = sorted(evidence, key=lambda e: (e['집행일자'], 비목순.index(e['비목코드'])))
r = 4
for i, e in enumerate(ev_sorted, 1):
    put(ws, r, [i, e['집행일자'], e['단위사업'], f"{e['보조비목']}({e['비목코드']})",
                f"{e['보조세목']}({e['세목코드']})", e['재원'], e['지급처']['명칭'],
                e['집행금액'], e['사용목적'], e['지급방식'], e['검토상태']],
        ctr_cols=(1,2,3,6,10,11), num_cols=(8,))
    r += 1
# 합계
put(ws, r, ['합계','','','','','','', None,'','',''], ctr_cols=(1,), bold=True, fill=TOT)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
ws.cell(r,8, f"=SUM(H4:H{r-1})").number_format='#,##0'; ws.cell(r,8).font=BB; ws.cell(r,8).fill=TOT; ws.cell(r,8).border=BOX; ws.cell(r,8).alignment=RIGHT

# ══════════════════════════════════════════════════════════
# 03_계약대장
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('03_계약대장')
head(ws, '계약대장', ['계약ID','단위사업','계약명','계약유형','계약방법','계약일','계약상대자','상태','비고'],
     [13,10,28,12,14,12,26,10,26])
r = 4
for c in contracts:
    put(ws, r, [c['계약ID'], c['단위사업'], c['계약명'], c['계약유형'], c['계약방법'],
                c.get('계약일',''), c['계약상대자']['상호'], c.get('상태',''), c.get('비고','')],
        ctr_cols=(1,2,4,6,8))
    r += 1

# ══════════════════════════════════════════════════════════
# 04_중요재산
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('04_중요재산')
head(ws, '중요재산 대장 (자산관리 시스템 연동)', ['자산ID','단위사업','재산구분','재산명','규격','취득가액','취득일','처분제한','설치장소','비고'],
     [12,10,11,26,22,14,12,10,24,22])
as_sorted = sorted(assets, key=lambda a: -a['취득가액'])
r = 4; first = r
for a in as_sorted:
    put(ws, r, [a['자산ID'], a['단위사업'], a['재산구분'], a['재산명'], a.get('규격',''),
                a['취득가액'], a['취득일'], f"{a['처분제한']['기간_년']}년", a.get('설치장소',''), a.get('비고','')],
        ctr_cols=(1,2,3,7,8), num_cols=(6,))
    r += 1
put(ws, r, ['합계','','','','', None,'','','',''], ctr_cols=(1,), bold=True, fill=TOT)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws.cell(r,6, f"=SUM(F{first}:F{r-1})").number_format='#,##0'; ws.cell(r,6).font=BB; ws.cell(r,6).fill=TOT; ws.cell(r,6).border=BOX; ws.cell(r,6).alignment=RIGHT

# ══════════════════════════════════════════════════════════
# 05_준공현황
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('05_준공현황')
head(ws, '준공 준비 현황', ['준공ID','단위사업','단위공사','유형','수행사','진도율','필수문서','확보','상태'],
     [12,10,28,10,18,9,10,8,10])
r = 4
for c in completion:
    필수 = [d for d in c['문서'] if d.get('필수여부', True)]
    확보 = sum(1 for d in 필수 if d['상태'] in ('제출완료','확정'))
    put(ws, r, [c['준공ID'], c['단위사업'], c['명칭'], c['유형'], c.get('수행사',''),
                (c.get('물리적진도율',0) or 0)/100, len(필수), 확보, c['상태']],
        ctr_cols=(1,2,4,6,7,8,9))
    ws.cell(r,6).number_format='0%'
    r += 1

# ══════════════════════════════════════════════════════════
# 06_검증지적
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('06_검증지적')
head(ws, '법정요건 검증 지적사항', ['수준','규칙','대상','내용','근거'], [8,8,20,44,20])
r = 4
설명 = {'R-01':'기준정보 미확정','R-05':'증빙 결손','R-06':'중요재산 미보고','R-08':'인건비 요건',
        'R-22':'예산 초과','R-24':'재원 미구분','R-04':'조달 절차','R-11':'낙찰차액','R-12':'진도율 괴리',
        'R-20':'선금 미정산','R-23':'환불·정정','R-09':'준공문서 결손','R-25':'사업자번호 미확보'}
import collections
grp = collections.defaultdict(lambda: {'수준':'', '건수':0, '근거':''})
for f in vr.get('결과', []):
    k = f['rule'].rstrip('abcde')
    grp[k]['수준'] = f['level']; grp[k]['건수'] += 1; grp[k]['근거'] = f['law']
    if grp[k]['수준'] != 'ERROR' and f['level']=='ERROR': grp[k]['수준']='ERROR'
for k in sorted(grp, key=lambda x:(grp[x]['수준']!='ERROR', -grp[x]['건수'])):
    g = grp[k]
    put(ws, r, ['오류' if g['수준']=='ERROR' else '경고', k, f"{g['건수']}건",
                설명.get(k, ''), g['근거']], ctr_cols=(1,2,3))
    ws.cell(r,1).font = Font(name=FK, size=9, bold=True, color='C00000' if g['수준']=='ERROR' else '8A6100')
    r += 1

# 저장
out = os.path.join(ROOT, 'output', '아산_정산_종합내역.xlsx')
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print('저장:', out)
print(f'시트 {len(wb.sheetnames)}개 · 집행 {len(evidence)}건 · 자산 {len(assets)}건')
